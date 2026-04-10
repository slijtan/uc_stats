#!/usr/bin/env python3
"""
Preprocess CRDC state file into a simple CSV for the pipeline.

The CRDC state XLSX file has no headers and 1700+ columns per school.
This script extracts only the fields we need and saves as a simple CSV.

Input:  raw-data/crdc/CRDC_2020-21_CA.xlsx
Output: raw-data/crdc/crdc_ap_courses.csv

Columns extracted:
  - COMBOKEY (col 6): 12-digit NCES ID (LEA ID + School ID)
  - SCH_APENR_IND (col 420): AP program indicator (Yes/No)
  - SCH_APCOURSES (col 421): Count of different AP courses offered

Column positions derived from CRDC 2020-21 Data Dictionary (Appendix B).
"""

import csv
import sys
import os

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# Column indices (0-based) in the CRDC state file
COL_COMBOKEY = 6
COL_SCHOOL_NAME = 5
COL_AP_INDICATOR = 420
COL_AP_COURSES = 421

def main():
    input_path = "raw-data/crdc/CRDC_2020-21_CA.xlsx"
    output_path = "raw-data/crdc/crdc_ap_courses.csv"

    if not os.path.exists(input_path):
        print(f"Error: Input file not found: {input_path}")
        sys.exit(1)

    print(f"Reading {input_path}...")
    wb = openpyxl.load_workbook(input_path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    schools_total = 0
    schools_with_ap = 0

    with open(output_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["combokey", "school_name", "ap_indicator", "ap_courses_offered"])

        for row in ws.iter_rows(values_only=True):
            vals = list(row)
            if len(vals) <= COL_AP_COURSES:
                continue

            combokey = str(vals[COL_COMBOKEY] or "").strip()
            school_name = str(vals[COL_SCHOOL_NAME] or "").strip()
            ap_indicator = str(vals[COL_AP_INDICATOR] or "").strip()
            ap_courses_raw = vals[COL_AP_COURSES]

            if not combokey:
                continue

            schools_total += 1

            # Parse AP courses count (-9 = not applicable, -13 = not available)
            ap_courses = None
            if isinstance(ap_courses_raw, (int, float)) and ap_courses_raw >= 0:
                ap_courses = int(ap_courses_raw)
            elif ap_indicator == "Yes" and (ap_courses_raw == -9 or ap_courses_raw is None):
                # School has AP program but course count not reported
                ap_courses = None

            if ap_indicator == "Yes":
                schools_with_ap += 1

            writer.writerow([
                combokey,
                school_name,
                ap_indicator,
                ap_courses if ap_courses is not None else "",
            ])

    wb.close()

    print(f"Processed {schools_total} schools")
    print(f"Schools with AP: {schools_with_ap}")
    print(f"Output: {output_path}")


if __name__ == "__main__":
    main()
